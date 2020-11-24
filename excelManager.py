from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
import traceback

class ExcelManager():
    alphabet = [
        'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
        'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'
    ]

    def __init__(self, fname, emp_list):
        super().__init__()
        self.FNAME = str(fname)
        self.EMP_LIST = emp_list
        self.WB = None
        self.WS = None

    # Excel File Management #
    def openExcel(self):
        """
        엑셀 파일을 열기 위한 메소드

        Returns:
            (Error Message)
        """
        try:
            self.WB = load_workbook(f'Excel/{self.FNAME}.xlsx')
            resultCheck = self.checkEmpList(self.WB.active)
            if resultCheck is not None:
                return resultCheck
        except Exception as e:
            return f'ExcelManager Error(openExcel()):\n{traceback.format_exc()}'

    def createExcel(self):
        """
        엑셀 파일을 생성하기 위한 메소드
        자동으로 'Statistics' 시트 생성

        Returns:
            (Error Message)
        """
        try:
            self.WB = Workbook()
            self.WB.active.title = 'Statistics'
            resultInit = self.initStatistics()
            if resultInit is not None:
                return resultInit
        except Exception as e:
            return f'ExcelManager Error(createExcel()):\n{traceback.format_exc()}'

    def saveExcel(self):
        """
        엑셀 파일을 저장하기 위한 메소드

        Returns:
            (Error Message)
        """
        try:
            self.WB.save(f'Excel/{self.FNAME}.xlsx')
        except Exception as e:
            return f'ExcelManager Error(saveExcel):\n{traceback.format_exc()}'

    # Worksheet Management #
    def getSheets(self):
        """
        엑셀 파일에 존재하는 시트 목록 반환하기 위한 메소드

        Returns:
            [T/F, [Worksheet Names]/'msg']
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')

            return [True, self.WB.sheetnames]
        except Exception as e:
            return [False, f'ExcelManager Error(getSheets()):\n{traceback.format_exc()}']

    def openSheet(self, ws_name=None):
        """
        ws_name에 해당하는 Worksheet를 열기 위한 메소드

        Parameters:
            ws_name: Worksheet 이름
        Returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if ws_name is None:
                raise Exception('Need to Worksheet Name(ws_name) Parameter')

            self.WS = self.WB[str(ws_name)]
            resultCheck = self.checkEmpList(self.WS)
            if resultCheck is not None:
                return resultCheck
        except Exception as e:
            return f'ExcelManager Error(openSheet):\n{traceback.format_exc()}'

    def createSheet(self, ws_name=None, last_day=None):
        """
        ws_name을 이름으로 하는 Worksheet를 생성하기 위한 메소드

        Parameters:
            ws_name: Worksheet 이름
            last_day: 마지막 날짜
        Returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if ws_name is None:
                raise Exception('Need to Worksheet Name(ws_name) Parameter')
            if last_day is None:
                raise Exception('Need to Last Day(last_day) Parameter')

            self.WS = self.WB.create_sheet()
            self.WS.title = str(ws_name)
            resultInit = self.initSheet(int(last_day))
            if resultInit is not None:
                return resultInit
            resultSave = self.saveExcel()
            if resultSave is not None:
                return resultSave
        except Exception as e:
            return f'ExcelManager Error(createSheet):\n{traceback.format_exc()}'

    def initStatistics(self):
        """
        엑셀 파일 생성 시 Statistics 시트를 초기화하기 위한 메소드

        Returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')

            ws = self.WB.active

            ## 이름, 근무지 입력
            ws.cell(row=3, column=2, value='이름').alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=3, column=3, value='근무지').alignment = Alignment(horizontal='center', vertical='center')
            for i, emp in enumerate(self.EMP_LIST):
                row = 3 * i + 4
                ws.merge_cells(start_row=row, end_row=row+2, start_column=2, end_column=2)
                ws.cell(row=row, column=2, value=emp).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=3, value='본사').alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row+1, column=3, value=400).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row+2, column=3, value='어비리').alignment = Alignment(horizontal='center', vertical='center')
            
            ws.freeze_panes = 'D4'
        except Exception as e:
            return f'ExcelManager Error(initStatistics()):\n{traceback.format_exc()}'

    def initSheet(self, last_day=None):
        """
        Worksheet 생성 시 템플릿에 맞춰 초기화하기 위한 메소드

        Returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if self.WS is None:
                raise Exception('Open Worksheet First')
            if last_day is None:
                raise Exception('Need to Last Day(last_day) Parameter')

            thin = Side(border_style='thin')

            ## 출퇴근 시각 칼럼 저장
            whTimeCells = [3 + 5 * n for n in range(last_day)]

            ## 이름, 근무지 입력
            self.WS.cell(row=3, column=2, value='이름').alignment = Alignment(horizontal='center', vertical='center')
            self.WS.cell(row=3, column=3, value='근무지').alignment = Alignment(horizontal='center', vertical='center')
            for i, emp in enumerate(self.EMP_LIST):
                row = 3 * i + 4
                self.WS.merge_cells(start_row=row, end_row=row+2, start_column=2, end_column=2)
                self.WS.cell(row=row, column=2, value=emp).alignment = Alignment(horizontal='center', vertical='center')
                self.WS.cell(row=row, column=3, value='본사').alignment = Alignment(horizontal='center', vertical='center')
                self.WS.cell(row=row+1, column=3, value=400).alignment = Alignment(horizontal='center', vertical='center')
                self.WS.cell(row=row+2, column=3, value='어비리').alignment = Alignment(horizontal='center', vertical='center')

            ## 셀 병합, 날짜 입력, 요일에 따른 배경 색 적용, 칼럼 입력
            columns = ['출퇴근 시각', '근무 시각', '근무 시간', '잔업 시간', '특근 시간']
            satBG = PatternFill(start_color='00BDD7EE', end_color='00BDD7EE', fill_type='solid')
            sunBG = PatternFill(start_color='00F8CBAD', end_color='00F8CBAD', fill_type='solid')
            for n in whTimeCells:
                day = whTimeCells.index(n) + 1
                n += 1
                import datetime
                weekday = datetime.date(int(self.FNAME), int(self.WS.title), day).weekday()

                self.WS.merge_cells(start_row=2, end_row=2, start_column=n, end_column=n + 4)
                self.WS.cell(row=2, column=n, value=day).alignment = Alignment(horizontal='center', vertical='center')
                if weekday == 5 or weekday == 6:
                    self.WS.cell(row=2, column=n).fill = satBG if weekday == 5 else sunBG
                for cell, column in zip(self.WS[3][n - 1:n + 4], columns):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.value = column
                
                for i, emp in enumerate(self.EMP_LIST):
                    self.WS.merge_cells(start_row=3 * i + 4, end_row=3 * i + 6, start_column=n, end_column=n)
                    
                for row in self.WS.iter_rows(min_row=2, max_row=self.WS.max_row, min_col=n+4, max_col=n+4):
                    for col in row:
                        col.border = Border(right=thin)

            ## 셀 너비 초기화
            for col in whTimeCells:
                self.WS.column_dimensions[ExcelManager.alphabet[(col // len(ExcelManager.alphabet)) - 1] + ExcelManager.alphabet[col % len(ExcelManager.alphabet)] if col // len(ExcelManager.alphabet) else ExcelManager.alphabet[col]].width = 12.5
                self.WS.column_dimensions[ExcelManager.alphabet[((col + 1) // len(ExcelManager.alphabet)) - 1] + ExcelManager.alphabet[(col + 1) % len(ExcelManager.alphabet)] if (col + 1) // len(ExcelManager.alphabet) else ExcelManager.alphabet[(col + 1)]].width = 12.5

            ## Statistics 시트에 칼럼 추가
            resultAdd = self.addMonthInStatistics()
            if resultAdd is not None:
                return resultAdd

            self.WS.freeze_panes = 'D4'
        except Exception as e:
            return f'ExcelManager Error(initSheet()):\n{traceback.format_exc()}'

    def addMonthInStatistics(self):
        """
        시트 생성 시 Statistics에 그 달에 해당하는 칼럼 추가하는 메소드

        Returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            
            month = int(self.WS.title)
            coln = month * 3 + 1
            ws = self.WB.active
            thin = Side(border_style='thin')

            ## 셀 병합, 월 입력, 칼럼 입력
            ws.merge_cells(start_row=2, end_row=2, start_column=coln, end_column=coln+2)
            ws.cell(row=2, column=coln, value=int(self.WS.title)).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=3, column=coln, value='근무 시간').alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=3, column=coln+1, value='잔업 시간').alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=3, column=coln+2, value='특근 시간').alignment = Alignment(horizontal='center', vertical='center')

            ## 수식 입력
            for row in ws.iter_rows(min_row=4, max_row=len(self.EMP_LIST)*3+3, min_col=coln, max_col=coln+2):
                for col in row:
                    col.alignment = Alignment(horizontal='center', vertical='center')
                    col.value = '=IFERROR(SUMIF(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$D$3:"&ADDRESS(3,COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$3:$3"))-2)),INDIRECT(ADDRESS(3,COLUMN())),OFFSET(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$A$1"),1+MATCH(INDIRECT(ADDRESS(ROW(),3)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$C$"&3+MATCH(INDIRECT(ADDRESS(IF(MOD(ROW(),3)=1,ROW(),IF(MOD(ROW(),3)=2,ROW()-1,ROW()-2)),2)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B$4:$B$"&3*COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B:$B"))),0)&":$C$"&5+MATCH(INDIRECT(ADDRESS(IF(MOD(ROW(),3)=1,ROW(),IF(MOD(ROW(),3)=2,ROW()-1,ROW()-2)),2)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B$4:$B$"&3*COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B:$B"))),0)))+MATCH(INDIRECT(ADDRESS(IF(MOD(ROW(),3)=1,ROW(),IF(MOD(ROW(),3)=2,ROW()-1,ROW()-2)),2)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B$4:$B$"&3*COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B:$B"))),0),3,1,COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$3:$3"))-2)), 0)'

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=coln+2, max_col=coln+2):
                for col in row:
                    col.border = Border(right=thin)
        except Exception as e:
            return f'ExcelManager Error(addMonthInStatistics()):\n{traceback.format_exc()}'

    # Data Management #
    def getData(self, name=None, day=None):
        """
        name, day와 일치하는 날짜에 해당하는 셀의 데이터를 조회하는 메소드

        Parameters:
            name: 조회할 직원
            day: 조회할 날짜
        Returns:
            [T/F, dataDict/'msg']
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if self.WS is None:
                raise Exception('Open Worksheet First')
            if name is None:
                raise Exception('Need to Employee Name(name) Parameter')
            if day is None:
                raise Exception('Need to Day(day) Parameter')

            indexDict = self.getIndex(name, day)
            if indexDict[0]:
                indexDict = indexDict[1]
            else:
                return indexDict

            dataDict = {
                'name': name,
                'date': {'y': self.FNAME, 'm': self.WS.title, 'd': day},
                'totalWorking': None,
                'hq': [None] * 4,
                '400': [None] * 4,
                'eobiri': [None] * 4
                }
            ## 출퇴근 시각
            dataDict['totalWorking'] = self.WS.cell(row=indexDict['name'], column=indexDict['day']).value
            ## 근무 데이터
            for row in self.WS.iter_rows(min_row=indexDict['name'], max_row=indexDict['name']+2, min_col=indexDict['day']+1, max_col=indexDict['day']+4):
                for col in row:
                    placeName = self.WS.cell(row=col.row, column=3).value
                    placeName = 'hq' if placeName == '본사' else '400' if placeName == 400 else 'eobiri' if placeName == '어비리' else None
                    columnName = self.WS.cell(row=3, column=col.column).value
                    columnName = 0 if columnName == '근무 시각' else 1 if columnName == '근무 시간' else 2 if columnName == '잔업 시간' else 3 if columnName == '특근 시간' else None
                    if placeName is None or columnName is None:
                        raise Exception('Not Exist Column')
                    dataDict[placeName][columnName] = col.value

            dataDict['hq'] = None if len(set(dataDict['hq'])) == 1 else dataDict['hq']
            dataDict['400'] = None if len(set(dataDict['400'])) == 1 else dataDict['400']
            dataDict['eobiri'] = None if len(set(dataDict['eobiri'])) == 1 else dataDict['eobiri']

            return [True, dataDict]
        except Exception as e:
            return [False, f'ExcelManager Error(getData()):\n{traceback.format_exc()}']

    def inputData(self, dataDict=None):
        """
        셀에 데이터를 입력하는 메소드

        Parameters:
            dataDict: 입력할 데이터 딕셔너리
        Returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if self.WS is None:
                raise Exception('Open Worksheet First')
            if dataDict is None:
                raise Exception('Need to Data Dictionary(dataDict) Parameter')

            indexDict = self.getIndex(dataDict['name'], dataDict['date']['d'])
            if indexDict[0]:
                indexDict = indexDict[1]
            else:
                return indexDict[1]

            ## 출퇴근 시각
            self.WS.cell(row=indexDict['name'], column=indexDict['day'], value=dataDict['totalWorking']).alignment = Alignment(horizontal='center', vertical='center')
            ## 근무 데이터
            for row in self.WS.iter_rows(min_row=indexDict['name'], max_row=indexDict['name']+2, min_col=indexDict['day']+1, max_col=indexDict['day']+4):
                for col in row:
                    placeName = self.WS.cell(row=col.row, column=3).value
                    placeName = 'hq' if placeName == '본사' else '400' if placeName == 400 else 'eobiri' if placeName == '어비리' else None
                    columnName = self.WS.cell(row=3, column=col.column).value
                    columnName = 0 if columnName == '근무 시각' else 1 if columnName == '근무 시간' else 2 if columnName == '잔업 시간' else 3 if columnName == '특근 시간' else None
                    if placeName is None or columnName is None:
                        raise Exception('Not Exist Column')
                    col.alignment = Alignment(horizontal='center', vertical='center')
                    col.value = dataDict[placeName][columnName] if dataDict[placeName] is not None and dataDict[placeName][columnName] is not None else ''
        except Exception as e:
            return f'ExcelManager Error(inputData()):\n{traceback.format_exc()}'

    # ETC #
    def checkEmpList(self, ws=None):
        """
        직원이 입력되어 있는지 확인하는 메소드
        
        Parameters:
            ws: 확인할 Worksheet
        Returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if ws is None:
                raise Exception('Need to Worksheet(ws) Parameter')

            ## 직원 수 비교
            if len(self.EMP_LIST) == ws.max_row / 3 - 1:
                return
            
            ## 입력된 직원 리스트 가져오기
            inputedEmpList = []
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=2):
                for col in row:
                    if col.row % 3 == 1:
                        inputedEmpList.append(col.value)

            ## 입력되어있지 않은 직원 거르기
            notInputedEmpList = [emp for emp in self.EMP_LIST if emp not in inputedEmpList]

            ## 직원 추가
            result = self.addEmp(ws, notInputedEmpList)
            if result is not None:
                return result
        except Exception as e:
            return f'ExcelManager Error(checkEmpList()):\n{traceback.format_exc()}'

    def addEmp(self, ws=None, emp_list=None):
        """
        ws 시트에 입력되어 있지 않은 직원들을 추가하는 메소드

        Parameters:
            ws: 추가할 Worksheet
            emp_list: 직원 목록
        returns:
            (Error Message)
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if ws is None:
                raise Exception('Need to Worksheet(ws) Parameter')
            if emp_list is None:
                raise Exception('Need to Employee List(emp_list) Parameter')

            thin = Side(border_style='thin')

            before_max_row = ws.max_row
            ## 시트 마지막에 직원 추가
            for i, emp in enumerate(emp_list):
                row = 3 * i + before_max_row + 1
                ws.merge_cells(start_row=row, end_row=row+2, start_column=2, end_column=2)
                ws.cell(row=row, column=2, value=emp).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=3, value='본사').alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row+1, column=3, value=400).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row+2, column=3, value='어비리').alignment = Alignment(horizontal='center', vertical='center')
            
            ## Statistics 시트일 경우 추가된 월 칼럼에 수식 추가
            if ws.title == 'Statistics':
                ### 입력되어있는 월 확인
                months = self.getSheets()
                if not months[0]:
                    return months[1]
                months = months[1]
                if 'Statistics' in months:
                    months.remove('Statistics')
                if len(months):
                    for month in months:
                        month = int(month)
                        coln = month * 3 + 1

                        #### 수식 입력
                        for row in ws.iter_rows(min_row=before_max_row+1, max_row=len(emp_list)*3+before_max_row, min_col=coln, max_col=coln+2):
                            for col in row:
                                col.alignment = Alignment(horizontal='center', vertical='center')
                                col.value = '=IFERROR(SUMIF(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$D$3:"&ADDRESS(3,COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$3:$3"))-2)),INDIRECT(ADDRESS(3,COLUMN())),OFFSET(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$A$1"),1+MATCH(INDIRECT(ADDRESS(ROW(),3)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$C$"&3+MATCH(INDIRECT(ADDRESS(IF(MOD(ROW(),3)=1,ROW(),IF(MOD(ROW(),3)=2,ROW()-1,ROW()-2)),2)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B$4:$B$"&3*COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B:$B"))),0)&":$C$"&5+MATCH(INDIRECT(ADDRESS(IF(MOD(ROW(),3)=1,ROW(),IF(MOD(ROW(),3)=2,ROW()-1,ROW()-2)),2)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B$4:$B$"&3*COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B:$B"))),0)))+MATCH(INDIRECT(ADDRESS(IF(MOD(ROW(),3)=1,ROW(),IF(MOD(ROW(),3)=2,ROW()-1,ROW()-2)),2)),INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B$4:$B$"&3*COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$B:$B"))),0),3,1,COUNTA(INDIRECT(INDIRECT(ADDRESS(2,IF(MOD(COLUMN(),3)=1,COLUMN(),IF(MOD(COLUMN(),3)=2,COLUMN()-1,COLUMN()-2))))&"!$3:$3"))-2)), 0)'

                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=coln+2, max_col=coln+2):
                            for col in row:
                                col.border = Border(right=thin)
            else:
                for i in range(len(emp_list)):
                    row = 3 * i + before_max_row + 1
                    for j in range(4, ws.max_column + 1, 5):
                        ws.merge_cells(start_row=row, end_row=row+2, start_column=j, end_column=j)
                        ws.cell(row=row, column=j).alignment = Alignment(horizontal='center', vertical='center')

                cols = [i for i in range(8, ws.max_column + 1, 5)]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=ws.max_column):
                    for col in row:
                        if col.column in cols:
                            col.border = Border(right=thin)
        except Exception as e:
            return f'ExcelManager Error(addEmp()):\n{traceback.format_exc()}'

    def getIndex(self, name=None, day=None):
        """
        name, day와 일치하는 날짜에 해당하는 셀의 데이터 인덱스를 조회하는 메소드

        Parameters:
            name: 조회할 직원
            day: 조회할 날짜
        Returns:
            [T/F, {'name': int, 'day': int}/'msg']
        """
        try:
            if self.WB is None:
                raise Exception('Open Excel First')
            if self.WS is None:
                raise Exception('Open Worksheet First')
            if name is None:
                raise Exception('Need to Employee Name(name) Parameter')
            if day is None:
                raise Exception('Need to Day(day) Parameter')

            indexDict = {'name': None, 'day': None}
            for row in self.WS.iter_rows(min_row=self.WS.min_row, max_row=self.WS.max_row, min_col=2, max_col=2):
                for col in row:
                    if col.value == name:
                        indexDict['name'] = col.row
                        break
            for col in self.WS.iter_cols(min_row=2, max_row=2, min_col=self.WS.min_column, max_col=self.WS.max_column):
                for row in col:
                    if str(row.value) == day:
                        indexDict['day'] = row.column
                        break

            return [True, indexDict]
        except Exception as e:
            return [False, f'ExcelManager Error(getIndex()):\n{traceback.format_exc()}']